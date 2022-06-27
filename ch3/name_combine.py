import openpyxl as excel
# 입출력 파일 지정
in_file = './input/name2.xlsx'
out_file = './output/name_combine.xlsx'

# 입력 문서를 열고 시트 가져오기 --- (1)
in_book = excel.load_workbook(in_file)
in_sheet = in_book.worksheets[0]

# 신규 문서를 생성해 시트 가져오기 --- (2)
out_book = excel.Workbook()
out_sheet = out_book.active

# 시트의 행 읽기 --- (3)
for row in in_sheet.iter_rows():
    # 성과 이름 가져오기 --- (3a)
    sung = row[0].value
    myung = row[1].value
    # 성과 이름 합치기 --- (3b)
    name = sung + ' ' + myung
    # 신규 시트에 추가 --- (3c)
    out_sheet.append([name])

# 결과 저장 --- (4)
out_book.save(out_file)
