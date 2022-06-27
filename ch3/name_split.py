import openpyxl as excel
# 입출력 파일 지정
in_file = './input/name1.xlsx'
out_file = './output/name_split.xlsx'

# 입력 문서를 열고 시트 가져오기 --- (1)
book = excel.load_workbook(in_file)
sheet = book.worksheets[0]

# 신규 문서를 생성해 시트 가져오기 --- (2)
out_book = excel.Workbook()
out_sheet = out_book.active

# 시트의 행 읽기 --- (3)
for row in sheet.iter_rows():
    # 앞뒤 공백이 제거된 성명 가져오기 --- (3a)
    name = row[0].value.strip()
    # 성과 이름 나누기 --- (3b)
    if ' ' in name: 
        sung, myung = name.split(' ')
    else : 
        sung, myung = name[0], name[1:] 
    # 신규 시트에 추가하기 --- (3c)
    out_sheet.append([sung, myung])
# 결과를 저장
out_book.save(out_file) 
