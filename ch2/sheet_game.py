import openpyxl as excel
import random

# 당첨 시트 번호 정하기 --- (1)
win_num = random.randint(1,100)

# 새 워크북 생성 --- (2)
book = excel.Workbook()
book.active["B2"] = "'당첨'이라고 적힌 시트를 찾아보자"

# 반복문을 돌며 100개의 시트 만들기 --- (3)
for i in range(1,101):
    # 새 워크시트 생성 --- (4)
    sname = str(i) + "번"
    sheet = book.create_sheet(title=sname)
    # 시트에 쓸 단어 정하기
    word = "꽝"
    if i == win_num: word = "당첨"
    # 임팩트가 있도록 화면을 word로 채우기 --- (5)
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y+1, x+1)
            c.value = word

# 파일 저장 --- (6)
book.save("./output/sheet_game.xlsx")
print("ok, winning number=", win_num)

