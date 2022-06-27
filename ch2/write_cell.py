# openpyxl 불러오기
import openpyxl as excel

# 워크북을 생성하고 활성화된 워크시트 가져오기
book = excel.Workbook()
sheet = book.active

# A1에 값 설정 --- (1)
sheet["A1"] = "일찍 일어나는 새가 벌레를 잡는다"

# A2(row=2, column=1)에 값 설정 --- (2)
sheet.cell(row=2, column=1, value="하늘은 스스로 돕는 자를 돕는다")

# A3(row=3, column=1)에 값 설정 --- (3)
third_cell = sheet.cell(row=3, column=1)
third_cell.value = "낙숫물이 바위를 뚫는다"

# 워크북 저장
book.save("output/write_cell.xlsx")
