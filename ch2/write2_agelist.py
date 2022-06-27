import openpyxl as excel
import datetime

# 새 워크북을 생성하고 워크시트 가져오기
book = excel.Workbook()
sheet = book.active

# 올해 연도 구하기 --- (1)
thisyear =  datetime.datetime.now().year

# 1행에 헤더 설정 --- (2)
sheet["A1"] = "출생 연도"
sheet["B1"] = "세는 나이"
sheet["C1"] = "만 나이 (생일 후)"
sheet["D1"] = "만 나이 (생일 전)"

# 셀의 너비 조정 --- (3)
sheet.column_dimensions['C'].width=15
sheet.column_dimensions['D'].width=15

# 생년·나이 연속 데이터 채우기 --- (4)
for i in range(80):
    # 설정할 값을 계산 --- (4a)
    birth_year = thisyear - i
    korean_age = thisyear - birth_year +1 #세는 나이는 1살부터 시작
    man_age = {'after_bday':korean_age-1, 'before_bday':korean_age-2}
    
    # 셀을 읽어 값을 설정하기 --- (4b)
    year_cell = sheet.cell(i+2, 1)
    year_cell.value = str(birth_year) + "년생"
    
    age_cell = sheet.cell(i+2, 2)
    age_cell.value = str(korean_age) + "세"
    
    age_cell = sheet.cell(i+2, 3)
    age_cell.value = "만 " + str(man_age['after_bday']) + "세"

    age_cell = sheet.cell(i+2, 4)
    age_cell.value = "만 " + str(man_age['before_bday']) + "세"

# 예외 값 처리 --- (5)
sheet["D2"]="-"

# 파일 저장 --- (6)
book.save("output/write2_agelist.xlsx")

