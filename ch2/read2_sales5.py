import openpyxl as excel

# 매출 명세 문서를 열고 시트 가져오기
sheet = excel.load_workbook(
    "input/monthly_sales.xlsx", data_only=True).active

# iter_rows를 사용하여 모든 데이터 가져오기 --- (1)
for row in sheet.iter_rows(min_row=3):
    values = [cell.value for cell in row]
    if values[0] is None: break
    print(values)
