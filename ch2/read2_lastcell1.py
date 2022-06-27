import openpyxl as excel

sheet =  excel.load_workbook("input/monthly_sales.xlsx").active
print((sheet.max_row, sheet.max_column))
