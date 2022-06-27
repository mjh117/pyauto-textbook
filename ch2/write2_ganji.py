import openpyxl as excel
import datetime 

# 표의 내용을 리스트로 나타내기 --- (1)
gan=['갑을병정무기경신임계','甲乙丙丁戊己庚辛壬癸', '청청적적황황백백흑흑']
ji=['자축인묘진사오미신유술해', '子丑寅卯辰巳午未申酉戌亥', ['쥐','소','호랑이','토끼','용','뱀','말','양','원숭이','닭','개','돼지']]

# 특정 연도에서 간지로 변환하는 함수를 정의 --- (2)
def year_to_ganji(year):
    # 나머지 연산을 통해 연도를 인덱스로 변환하기 --- (2a)
    i = (year-4)%10
    j = (year-4)%12
    
    # 인덱스로 간지 정보 조회하기 --- (2b)
    ganji = gan[0][i]+ji[0][j]+'('+gan[1][i]+ji[1][j]+')'

    # 인덱스로 동물 정보 조회하기 --- (2c)
    color_animal = gan[2][i]+'색'+ji[2][j]
    return ganji, color_animal

# 새 워크북을 만들고 시트 가져오기 --- (3)
book = excel.Workbook()
sheet = book.active

# 헤더 설정하기 --- (4)
sheet["A1"] = "연도"
sheet["B1"] = "간지"
sheet["C1"] = "동물"

# 시작 연도를 올해로 설정하기 --- (5)
start_y = now = datetime.datetime.now().year

# 100년간의 연도·간지 정보를 시트에 채우기 --- (6)
for i in range(100):
    # 연도를 간지로 변환하기 --- (6a)
    year = start_y - i
    result = year_to_ganji(year)
    ganji = result[0]
    col_ani = result[1]
    # 시트에 입력하기 --- (6b)
    sheet.cell(row=(2+i), column=1, value=str(year)+'년')
    sheet.cell(row=(2+i), column=2, value=ganji)
    sheet.cell(row=(2+i), column=3, value=col_ani)
    print(year,"=", ganji, ",", col_ani) #IDLE 쉘 창에 연도·간지 정보 출력

# 파일 저장 --- (7)
book.save("output/write2_ganji.xlsx")
