import openpyxl as excel
import datetime

# 새 워크북을 생성하고 워크시트 가져오기
book = excel.Workbook()
sheet = book.active

# 헤더에 값 설정 --- (1)
sheet["A1"] = "출생 기간"
sheet["B1"] = "초등학교 입학 연도"
sheet["C1"] = "대학교 학번"

#셀의 너비 조정 --- (2)
sheet.column_dimensions['A'].width=40
sheet.column_dimensions['B'].width=20
sheet.column_dimensions['C'].width=20

# 셀에 연속 데이터 채우기 --- (3)
for i in range(50):
    # 기준 출생 연도 --- (3a)
    birth_year = 2002 - i

    #출생 기간 문자열 설정 --- (3b)
    birth_range = "{}년 3월 1일생 ~ {}년 2월 28(29)일생".format(birth_year, birth_year+1)

    #초등학교 입학 연도 계산 --- (3c)
    ele_year = birth_year + 7

    #대학교 학번 계산 --- (3d)
    univ_year = birth_year + 19
    univ_num = str(univ_year)[2:]

    # 셀을 지정하여 값을 설정하기 --- (3e)
    sheet.cell(i+2, 1, value= birth_range)
    sheet.cell(i+2, 2, value= str(ele_year) +"년")
    sheet.cell(i+2, 3, value= univ_num+"학번")

# 예외 값 처리 --- (4)
sheet["A2"]="2002년 3월 1일생 ~ 2002년 12월 31일생"

# 파일 저장
book.save("output/write2_entry_year.xlsx")
