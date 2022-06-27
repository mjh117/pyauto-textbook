import openpyxl as xl

book = xl.Workbook()
sheet = book.active

# 셀의 너비 설정 --- (1)
sheet.column_dimensions['B'].width = 40
# 셀의 높이 설정 --- (2)
sheet.row_dimensions[2].height = 40 

cell = sheet["B2"]
cell.value = "웃음은 만병의 통치약이다"

# 텍스트 배치 설정 --- (3)
from openpyxl.styles.alignment import Alignment
cell.alignment = Alignment(
        horizontal='center', # 수평 위치를 가운데로
        vertical='center') # 수직 위치를 가운데로

# 테두리 설정 --- (4)
from openpyxl.styles.borders import Border, Side
cell.border = Border(
    top=Side(style='thin', color='000000'),   # 위쪽
    right=Side(style='thin', color='000000'), # 오른쪽
    bottom=Side(style='thin', color='000000'),# 아래쪽
    left=Side(style='thin', color='000000'),  # 왼쪽
)

# 글꼴 설정 --- (5)
from openpyxl.styles import Font
cell.font = Font(
    size=14,        #크기
    bold=True,      # 굵게
    italic=True,    # 기울임꼴
    color='FFFFFF') # 색상

# 배경색 설정 --- (6)
from openpyxl.styles import PatternFill
cell.fill = PatternFill(
        fill_type='solid', # 전면 채우기 
        fgColor='FF0000')  # 빨간색

# 파일 저장
book.save("output/cellformat_style.xlsx")
